"""Render lookup results into a card-show spreadsheet."""

from __future__ import annotations

import io
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import branding, palette
from .export_fields import (
    ADJUSTED_COMP_80,
    ADJUSTED_COMP_85,
    ADJUSTED_COMP_90,
    ADJUSTED_COMP_95,
    ADJUSTED_MARKET,
    COMP_80,
    COMP_85,
    COMP_90,
    COMP_95,
    CONDITION,
    MARKET,
    NAME,
    NUMBER,
    RARITY,
    SET,
    SOURCE,
    SOURCE_URL,
    THUMBNAIL,
    VARIANT,
)
from .images import THUMB_H, THUMB_W, make_thumbnail
from .parser import CardQuery
from .pricing import COMP_PERCENTS, Pricing


@dataclass(frozen=True)
class _ColumnSpec:
    id: str
    field: str | None  # toggle key this column answers to; None = always on
    header: str
    width: int


# Column order as it appears left-to-right in the sheet. `field=None` marks
# columns that aren't part of the configurable-fields toggle (#262) and
# always render regardless of what a caller passes for `fields`.
_COLUMNS: tuple[_ColumnSpec, ...] = (
    _ColumnSpec("thumbnail", THUMBNAIL, "Image", 16),
    _ColumnSpec("src_tag", None, "Source", 14),  # the input file's tag
    _ColumnSpec("input", None, "Input", 28),
    _ColumnSpec("name", NAME, "Name", 24),
    _ColumnSpec("set", SET, "Set", 22),
    _ColumnSpec("series", None, "Series", 18),
    _ColumnSpec("number", NUMBER, "Number", 10),
    _ColumnSpec("rarity", RARITY, "Rarity", 14),
    _ColumnSpec("variant", VARIANT, "Variant", 16),
    _ColumnSpec("database", None, "Database", 18),
    _ColumnSpec("condition", CONDITION, "Condition", 12),
    _ColumnSpec("market", MARKET, "Market", 12),
    _ColumnSpec("adjusted_market", ADJUSTED_MARKET, "Adjusted Market", 16),
    _ColumnSpec("comp_80", COMP_80, "80%", 10),
    _ColumnSpec("comp_85", COMP_85, "85%", 10),
    _ColumnSpec("comp_90", COMP_90, "90%", 10),
    _ColumnSpec("comp_95", COMP_95, "95%", 10),
    _ColumnSpec("adjusted_comp_80", ADJUSTED_COMP_80, "Adjusted 80%", 14),
    _ColumnSpec("adjusted_comp_85", ADJUSTED_COMP_85, "Adjusted 85%", 14),
    _ColumnSpec("adjusted_comp_90", ADJUSTED_COMP_90, "Adjusted 90%", 14),
    _ColumnSpec("adjusted_comp_95", ADJUSTED_COMP_95, "Adjusted 95%", 14),
    _ColumnSpec("ebay_sold", None, "eBay Sold (median)", 16),
    _ColumnSpec("ebay_active", None, "eBay Active (floor)", 16),
    _ColumnSpec("source", SOURCE, "Price Source", 14),
    _ColumnSpec("source_url", SOURCE_URL, "Listing URL", 38),
)

# Preserved for callers that just want the full default header list (tests,
# docs) — always the unfiltered set, in sheet order.
HEADERS = [c.header for c in _COLUMNS]


def _active_columns(fields: frozenset[str] | None) -> list[_ColumnSpec]:
    if fields is None:
        return list(_COLUMNS)
    return [c for c in _COLUMNS if c.field is None or c.field in fields]


@dataclass
class Row:
    query: CardQuery
    card: dict[str, Any] | None
    pricing: Pricing
    image_path: Path | None = None
    tag: str = ""  # input-file stem so rows stay grouped per source list


# Rarity tiers the palette already defines (design/tokens/colors_and_type.css
# → palette.py). Bucketed by substring rather than an exact-match table since
# pokemontcg.io's rarity strings multiply constantly (Illustration Rare, Ultra
# Rare, Rare Secret, ACE SPEC Rare, …) — anything beyond plain Common/Uncommon
# reads as the same "rare" tier here. Order matters: "uncommon" must be
# checked before "common" since it contains that substring.
_RARITY_TIER_TOKENS: dict[str, str] = {
    "uncommon": "rarity-uncommon",
    "common": "rarity-common",
    "rare": "rarity-rare",
}


def _rarity_fill_token(rarity: str | None) -> str | None:
    """Return the palette token for `rarity`'s tier, or `None` if unmapped."""
    if not rarity:
        return None
    lowered = rarity.lower()
    for substring, token in _RARITY_TIER_TOKENS.items():
        if substring in lowered:
            return token
    return None


def _money_format(currency: str) -> str:
    if currency == "EUR":
        return '"€"#,##0.00'
    if currency == "GBP":
        return '"£"#,##0.00'
    return '"$"#,##0.00'


def write_spreadsheet(
    rows: list[Row],
    out_path: Path,
    max_price: float | None = None,
    fields: frozenset[str] | None = None,
) -> None:
    """Render `rows` into an xlsx workbook.

    `fields` restricts which of the configurable columns (#262) appear —
    `None` (the CLI default) renders every column, matching pre-#262
    behavior. Columns outside the toggle set (Source tag, Input, Series,
    Database, eBay comps) always render."""
    columns = _active_columns(fields)
    col_idx = {c.id: i for i, c in enumerate(columns, start=1)}

    wb = Workbook()
    _apply_workbook_branding(wb, out_path)
    ws = wb.active
    ws.title = "Cards"

    over_cap_fill = PatternFill("solid", fgColor=palette.hex("warning-bg"))  # above-cap rows
    rarity_fills = {
        token: PatternFill("solid", fgColor=palette.hex(token))
        for token in set(_RARITY_TIER_TOKENS.values())
    }

    _write_header_row(ws, columns)
    _apply_column_widths(ws, columns)

    for i, row in enumerate(rows, start=2):
        _write_data_row(ws, i, row, max_price, over_cap_fill, rarity_fills, col_idx)

    if "src_tag" in col_idx:
        freeze_col = get_column_letter(col_idx["src_tag"] + 1)
        ws.freeze_panes = f"{freeze_col}2"
    _write_totals_footer(ws, len(rows), col_idx)
    _apply_dark_body(ws, last_row=len(rows) + 3, ncols=len(columns))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _write_header_row(ws: Any, columns: list[_ColumnSpec]) -> None:
    header_font = Font(bold=True, color=palette.hex("fg-on-dark"))
    header_fill = PatternFill("solid", fgColor=palette.hex(palette.HEADER_BAND))
    ws.append([c.header for c in columns])
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_column_widths(ws: Any, columns: list[_ColumnSpec]) -> None:
    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = c.width
    ws.row_dimensions[1].height = 22


def _write_data_row(
    ws: Any,
    i: int,
    row: Row,
    max_price: float | None,
    over_cap_fill: Any,
    rarity_fills: dict[str, Any],
    col_idx: dict[str, int],
) -> None:
    ws.row_dimensions[i].height = THUMB_H * 0.78  # excel "points", approx px*0.75

    card = row.card or {}
    card_set = card.get("set") or {}
    money_fmt = _money_format(row.pricing.currency)

    _write_card_cells(ws, i, row, card, card_set, rarity_fills, col_idx)
    _write_pricing_cells(ws, i, row, money_fmt, max_price, over_cap_fill, col_idx)
    _embed_thumbnail(ws, i, row, col_idx)


def _write_card_cells(
    ws: Any,
    i: int,
    row: Row,
    card: dict[str, Any],
    card_set: dict,
    rarity_fills: dict[str, Any],
    col_idx: dict[str, int],
) -> None:
    rarity = card.get("rarity")
    values = {
        "src_tag": row.tag,
        "input": row.query.raw,
        "name": card.get("name") or "(not found)",
        "set": card_set.get("name"),
        "series": card_set.get("series"),
        "number": card.get("number"),
        "rarity": rarity,
        "variant": row.pricing.variant,
        "database": card.get("_database") or "",
    }
    for col_id, value in values.items():
        if col_id in col_idx:
            ws.cell(row=i, column=col_idx[col_id], value=value)

    if "rarity" in col_idx:
        fill_token = _rarity_fill_token(rarity)
        if fill_token is not None:
            cell = ws.cell(row=i, column=col_idx["rarity"])
            cell.fill = rarity_fills[fill_token]
            # Rarity tokens don't have dark-theme overrides (same bright fill
            # in both themes), so pin an explicit dark, on-brand-surface font
            # color rather than leaving it unset — otherwise `_apply_dark_body`
            # rewrites it to the light dark-theme foreground, which is nearly
            # unreadable against a light fill like rarity-rare/sun-300.
            cell.font = Font(color=palette.hex("fg-on-primary"))


def _write_pricing_cells(
    ws: Any,
    i: int,
    row: Row,
    money_fmt: str,
    max_price: float | None,
    over_cap_fill: Any,
    col_idx: dict[str, int],
) -> None:
    market = row.pricing.market_or_override
    adjusted_market = row.pricing.adjusted_market

    if "condition" in col_idx:
        ws.cell(row=i, column=col_idx["condition"], value=row.pricing.condition or "")

    _write_money_cell(ws, i, "market", market, money_fmt, max_price, over_cap_fill, col_idx)
    _write_money_cell(
        ws,
        i,
        "adjusted_market",
        adjusted_market,
        money_fmt,
        max_price,
        over_cap_fill,
        col_idx,
    )
    _write_comp_cells(
        ws,
        i,
        ("comp_80", "comp_85", "comp_90", "comp_95"),
        market,
        money_fmt,
        max_price,
        over_cap_fill,
        col_idx,
    )
    _write_comp_cells(
        ws,
        i,
        ("adjusted_comp_80", "adjusted_comp_85", "adjusted_comp_90", "adjusted_comp_95"),
        adjusted_market,
        money_fmt,
        max_price,
        over_cap_fill,
        col_idx,
    )
    _write_ebay_cells(ws, i, row, money_fmt, col_idx)
    _write_source_cells(ws, i, row, col_idx)


def _write_money_cell(
    ws: Any,
    i: int,
    col_id: str,
    value: float | None,
    money_fmt: str,
    max_price: float | None,
    over_cap_fill: Any,
    col_idx: dict[str, int],
) -> None:
    if col_id not in col_idx:
        return
    if value is None:
        ws.cell(row=i, column=col_idx[col_id], value="—")
        return

    cell = ws.cell(row=i, column=col_idx[col_id], value=value)
    cell.number_format = money_fmt
    if _is_over_cap(value, max_price):
        cell.fill = over_cap_fill
        cell.font = Font(bold=True, color=palette.hex("warning-fg"))
    else:
        # In-budget market in brand green, matching the binder/checklist.
        cell.font = Font(bold=True, color=palette.hex("success-fg"))


def _write_comp_cells(
    ws: Any,
    i: int,
    col_ids: tuple[str, str, str, str],
    market: float | None,
    money_fmt: str,
    max_price: float | None,
    over_cap_fill: Any,
    col_idx: dict[str, int],
) -> None:
    if market is None:
        return
    is_over_cap = _is_over_cap(market, max_price)
    for col_id, pct in zip(col_ids, COMP_PERCENTS, strict=True):
        if col_id not in col_idx:
            continue
        cell = ws.cell(row=i, column=col_idx[col_id], value=round(market * pct / 100, 2))
        cell.number_format = money_fmt
        if is_over_cap:
            cell.fill = over_cap_fill


def _write_ebay_cells(ws: Any, i: int, row: Row, money_fmt: str, col_idx: dict[str, int]) -> None:
    for col_id, value in (
        ("ebay_sold", row.pricing.ebay_sold_median),
        ("ebay_active", row.pricing.ebay_active_floor),
    ):
        if col_id not in col_idx:
            continue
        cell = ws.cell(row=i, column=col_idx[col_id], value=value if value is not None else "—")
        if value is not None:
            cell.number_format = money_fmt


def _write_source_cells(ws: Any, i: int, row: Row, col_idx: dict[str, int]) -> None:
    if "source" in col_idx:
        ws.cell(row=i, column=col_idx["source"], value=row.pricing.source or "")
    if "source_url" in col_idx and row.pricing.url:
        link_cell = ws.cell(row=i, column=col_idx["source_url"], value=row.pricing.url)
        link_cell.hyperlink = row.pricing.url
        link_cell.font = Font(color=palette.hex("fg-link"), underline="single")


def _is_over_cap(value: float, max_price: float | None) -> bool:
    return max_price is not None and value > max_price


def _embed_thumbnail(ws: Any, i: int, row: Row, col_idx: dict[str, int]) -> None:
    if "thumbnail" not in col_idx:
        return
    if not (row.image_path and row.image_path.exists()):
        return
    col_letter = get_column_letter(col_idx["thumbnail"])
    try:
        thumb_bytes = make_thumbnail(row.image_path, (THUMB_W, THUMB_H))
        xl_img = XLImage(io.BytesIO(thumb_bytes))
        xl_img.width = THUMB_W
        xl_img.height = THUMB_H
        anchor_cell = f"{col_letter}{i}"
        ws.add_image(xl_img, anchor_cell)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 16, THUMB_W / 7
        )
        ws.row_dimensions[i].height = THUMB_H * 0.78
    except Exception as exc:
        print(f"  ! thumbnail embed failed for {row.query}: {exc}", file=sys.stderr)


def _write_totals_footer(ws: Any, row_count: int, col_idx: dict[str, int]) -> None:
    # Summary footer. Note: SUM aggregates all rows regardless of currency, so
    # mixed-currency runs will produce an arithmetic-but-not-meaningful total.
    last = row_count + 2
    pricing_cols = [
        col_idx[col_id]
        for col_id in (
            "market",
            "adjusted_market",
            "comp_80",
            "comp_85",
            "comp_90",
            "comp_95",
            "adjusted_comp_80",
            "adjusted_comp_85",
            "adjusted_comp_90",
            "adjusted_comp_95",
        )
        if col_id in col_idx
    ]
    if "database" in col_idx:
        ws.cell(row=last + 1, column=col_idx["database"], value="Totals:").font = Font(bold=True)
    for col in pricing_cols:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{last - 1})"
        cell = ws.cell(row=last + 1, column=col, value=formula)
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(bold=True)

    # Branding mark in the totals row — clickable link back to the project
    # site so a recipient can trace where the file came from.
    if "src_tag" in col_idx:
        brand_cell = ws.cell(row=last + 1, column=col_idx["src_tag"], value=branding.PROJECT_NAME)
        brand_cell.hyperlink = branding.PROJECT_URL
        brand_cell.font = Font(bold=True, color=palette.hex("fg-link"), underline="single")


def _apply_dark_body(ws: Any, last_row: int, ncols: int) -> None:
    """Paint the data area for the dark theme (no-op in light).

    openpyxl has no sheet-level background — an unpainted cell stays white,
    which would wash out the dark palette's light `fg-1`/`success-fg`
    accents. Cells that already carry a themed fill (over-cap warning) or an
    explicit font color (market, links) keep them; everything else gets the
    dark surface + foreground."""
    if palette.current_theme() != "dark":
        return
    body_fill = PatternFill("solid", fgColor=palette.hex("bg-surface"))
    fg = palette.hex("fg-1")
    for row_cells in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=ncols):
        for cell in row_cells:
            if cell.fill.patternType is None:
                cell.fill = body_fill
            # The default font carries a *theme* color, not an rgb one — only
            # cells given an explicit rgb (market, links) keep their color.
            color = cell.font.color
            if color is None or color.type != "rgb":
                cell.font = Font(bold=cell.font.bold, underline=cell.font.underline, color=fg)


def _apply_workbook_branding(wb: Workbook, out_path: Path) -> None:
    """Stamp workbook properties + drop the logo into the top-left.

    Workbook properties land in the "Get Info" / "Properties" view of
    Excel and the OS so a recipient can see at a glance where the file
    came from. The logo image is anchored to A1 sized to fit inside the
    header row's existing height — no layout impact on the data rows."""
    props = wb.properties
    props.creator = branding.PROJECT_AUTHOR
    props.lastModifiedBy = branding.PROJECT_AUTHOR
    props.title = out_path.stem
    props.subject = f"Generated by {branding.PROJECT_NAME}"
    props.description = f"{branding.PROJECT_NAME} — {branding.PROJECT_URL}"
    props.keywords = "pokemon, tcg, card-collection, mgz-pkmn"

    ws = wb.active
    try:
        logo_data = branding.logo_bytes("on-dark")
    except (FileNotFoundError, OSError, ModuleNotFoundError) as exc:
        print(f"  ! branding logo unavailable, skipping xlsx logo: {exc}", file=sys.stderr)
        return
    try:
        logo = XLImage(io.BytesIO(logo_data))
        # Header row is 22 pt ≈ 29 px; scale the wordmark to ~24 px tall so it
        # sits flush inside the row without crowding the first column header.
        target_h = 24
        logo.height = target_h
        logo.width = int(target_h * branding.LOGO_ASPECT)
        ws.add_image(logo, "A1")
    except OSError as exc:
        # PIL / openpyxl couldn't decode the PNG bytes.
        print(f"  ! xlsx logo embed failed: {exc}", file=sys.stderr)
        return
