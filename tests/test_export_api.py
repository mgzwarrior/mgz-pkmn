from __future__ import annotations

import sys
import unittest
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fastapi.testclient import TestClient
from PIL import Image as PILImage

from api.main import app

client = TestClient(app)


def _row_payload(
    *,
    cid: str = "sv8-1",
    number: str = "1",
    name: str = "Pikachu",
    set_name: str = "Surging Sparks",
    market: float | None = 12.5,
    tag: str = "test",
    images: dict | None = None,
) -> dict:
    card = {
        "id": cid,
        "name": name,
        "number": number,
        "set": {"id": "sv8", "name": set_name, "printedTotal": 191, "total": 252},
        "_database": "pokemontcg.io",
        "language": "en",
    }
    if images is not None:
        card["images"] = images
    pricing = {"market": market, "currency": "USD"}
    return {
        "query": {"raw": name, "name": name},
        "card": card,
        "pricing": pricing,
        "tag": tag,
    }


def _fake_download(url, dest, session) -> bool:
    """Stand-in for images.download_image — writes a tiny valid PNG."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    PILImage.new("RGB", (10, 14), "white").save(dest, format="PNG")
    return True


class ExportApiTests(unittest.TestCase):
    def test_xlsx_export(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertGreater(len(resp.content), 0)

    def test_standard_pdf_export(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "pdf"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.headers["content-type"], "application/pdf")
        self.assertIn("binder.pdf", resp.headers["content-disposition"])

    def test_condensed_pdf_export(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "condensed-pdf"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.headers["content-type"], "application/pdf")
        self.assertIn("binder-condensed.pdf", resp.headers["content-disposition"])

    def test_checklist_export(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "checklist"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.headers["content-type"], "application/pdf")
        self.assertIn("checklist.pdf", resp.headers["content-disposition"])

    def test_checklist_with_no_matches_rejected(self) -> None:
        unmatched = {
            "query": {"raw": "Foo", "name": "Foo"},
            "card": None,
            "pricing": {"market": None, "currency": "USD"},
            "tag": "t",
        }
        resp = client.post(
            "/api/v1/export",
            json={"rows": [unmatched], "format": "checklist"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_unknown_format_rejected(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "bogus"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_unknown_sort_rejected(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx", "sort": "bogus"},
        )
        self.assertEqual(resp.status_code, 400)

    def test_xlsx_export_skips_images_by_default(self) -> None:
        payload = _row_payload(images={"large": "https://img.example/pikachu.png"})
        with patch("api.routes.export.download_image") as mock_dl:
            resp = client.post(
                "/api/v1/export",
                json={"rows": [payload], "format": "xlsx"},
            )
        self.assertEqual(resp.status_code, 200)
        mock_dl.assert_not_called()

    def test_xlsx_export_embeds_images_when_enabled(self) -> None:
        payload = _row_payload(images={"large": "https://img.example/pikachu.png"})
        with patch("api.routes.export.download_image", side_effect=_fake_download) as mock_dl:
            resp = client.post(
                "/api/v1/export",
                json={"rows": [payload], "format": "xlsx", "no_images": False},
            )
        self.assertEqual(resp.status_code, 200)
        self.assertGreater(len(resp.content), 0)
        mock_dl.assert_called_once()

    def test_xlsx_export_skips_download_when_thumbnail_field_disabled(self) -> None:
        # Images enabled, but the Thumbnail column itself is toggled off —
        # nothing renders it, so paying for the download is wasted work.
        payload = _row_payload(images={"large": "https://img.example/pikachu.png"})
        with patch("api.routes.export.download_image") as mock_dl:
            resp = client.post(
                "/api/v1/export",
                json={
                    "rows": [payload],
                    "format": "xlsx",
                    "no_images": False,
                    "fields": ["name", "market"],
                },
            )
        self.assertEqual(resp.status_code, 200)
        mock_dl.assert_not_called()

    def test_checklist_export_never_downloads_images(self) -> None:
        # The checklist is text-only — even with images enabled, no downloads.
        payload = _row_payload(images={"large": "https://img.example/pikachu.png"})
        with patch("api.routes.export.download_image") as mock_dl:
            resp = client.post(
                "/api/v1/export",
                json={"rows": [payload], "format": "checklist", "no_images": False},
            )
        self.assertEqual(resp.status_code, 200)
        mock_dl.assert_not_called()

    def test_xlsx_export_honors_fields_subset(self) -> None:
        from openpyxl import load_workbook

        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx", "fields": ["name"]},
        )
        self.assertEqual(resp.status_code, 200)
        import io

        ws = load_workbook(io.BytesIO(resp.content)).active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Name", header_row)
        self.assertNotIn("Rarity", header_row)
        self.assertNotIn("Market", header_row)

    def test_xlsx_export_carries_adjusted_condition_values(self) -> None:
        from openpyxl import load_workbook

        payload = _row_payload(market=100.0)
        payload["pricing"] |= {
            "condition": "LP",
            "condition_multiplier": 0.85,
            "adjusted_market": 85.0,
        }
        resp = client.post(
            "/api/v1/export",
            json={
                "rows": [payload],
                "format": "xlsx",
                "fields": ["condition", "market", "adjusted_market", "comp_80", "adjusted_comp_80"],
            },
        )
        self.assertEqual(resp.status_code, 200)
        import io

        ws = load_workbook(io.BytesIO(resp.content)).active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(ws.cell(row=2, column=header_row.index("Condition") + 1).value, "LP")
        self.assertEqual(ws.cell(row=2, column=header_row.index("Market") + 1).value, 100.0)
        self.assertEqual(
            ws.cell(row=2, column=header_row.index("Adjusted Market") + 1).value,
            85.0,
        )
        self.assertEqual(ws.cell(row=2, column=header_row.index("80%") + 1).value, 80.0)
        self.assertEqual(
            ws.cell(row=2, column=header_row.index("Adjusted 80%") + 1).value,
            68.0,
        )

    def test_unsupported_field_for_format_is_silently_dropped(self) -> None:
        # "rarity" isn't in the binder PDF's supported set — shouldn't 400.
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "pdf", "fields": ["name", "rarity"]},
        )
        self.assertEqual(resp.status_code, 200)

    def test_sort_accepts_documented_modes(self) -> None:
        for mode in ("number", "number-desc", "price-asc", "price-desc", "release-date", "alpha"):
            resp = client.post(
                "/api/v1/export",
                json={"rows": [_row_payload()], "format": "xlsx", "sort": mode},
            )
            self.assertEqual(resp.status_code, 200, f"sort={mode} failed: {resp.text}")


class ExportThemeTests(unittest.TestCase):
    """The `theme` field (#598) — validation plus proof it reaches the writer."""

    def test_unknown_theme_rejected(self) -> None:
        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx", "theme": "sepia"},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("theme", resp.json()["detail"])

    def test_dark_theme_accepted_for_every_format(self) -> None:
        for fmt in ("xlsx", "pdf", "condensed-pdf", "checklist"):
            resp = client.post(
                "/api/v1/export",
                json={"rows": [_row_payload()], "format": fmt, "theme": "dark"},
            )
            self.assertEqual(resp.status_code, 200, f"format={fmt} failed: {resp.text}")

    def test_dark_xlsx_paints_the_body(self) -> None:
        import io

        from openpyxl import load_workbook

        from mgz_pkmn import palette

        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx", "theme": "dark"},
        )
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(io.BytesIO(resp.content)).active
        name_cell = ws.cell(row=2, column=3)  # Source tag, Input, Name
        with palette.use_theme("dark"):
            dark_surface = palette.hex("bg-surface")
            dark_fg = palette.hex("fg-1")
        self.assertEqual(str(name_cell.fill.fgColor.rgb)[-6:], dark_surface)
        self.assertEqual(str(name_cell.font.color.rgb)[-6:], dark_fg)

    def test_light_xlsx_body_stays_unpainted(self) -> None:
        import io

        from openpyxl import load_workbook

        resp = client.post(
            "/api/v1/export",
            json={"rows": [_row_payload()], "format": "xlsx"},
        )
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(io.BytesIO(resp.content)).active
        name_cell = ws.cell(row=2, column=3)
        self.assertIsNone(name_cell.fill.patternType)


if __name__ == "__main__":
    unittest.main()
