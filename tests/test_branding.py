"""End-to-end branding checks for every export format.

Each format must:
* embed the project name + URL into file-properties metadata, and
* render the per-page footer (PDFs) or summary-row mark (xlsx) so a
  recipient can trace the file back to mgz-pkmn.

We only assert text presence and metadata-field population; pixel
positions are intentionally out of scope so layout tweaks don't churn
these tests.
"""

from __future__ import annotations

import re
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

# Disable ReportLab page compression for this test module so footer text
# stays as literal `(generated YYYY-MM-DD)` runs in the PDF bytes — the
# alternative is round-tripping every content stream through
# ASCII85 + zlib which is more machinery than these checks need.
import reportlab.rl_config

reportlab.rl_config.pageCompression = 0

from openpyxl import load_workbook  # noqa: E402

from mgz_pkmn import branding  # noqa: E402
from mgz_pkmn.binder import CONDENSED_LAYOUT, STANDARD_LAYOUT, write_binder_pdf  # noqa: E402
from mgz_pkmn.checklist import write_checklist_pdf  # noqa: E402
from mgz_pkmn.parser import CardQuery  # noqa: E402
from mgz_pkmn.pricing import Pricing  # noqa: E402
from mgz_pkmn.set_cards import write_set_cards_pdf  # noqa: E402
from mgz_pkmn.spreadsheet import Row, write_spreadsheet  # noqa: E402


def _row(tag: str = "list-a") -> Row:
    return Row(
        query=CardQuery(raw="Pikachu", name="Pikachu"),
        card={
            "id": "x",
            "name": "Pikachu",
            "number": "1",
            "set": {"name": "Base", "printedTotal": 102},
            "language": "en",
        },
        pricing=Pricing(market=12.5, currency="USD"),
        image_path=None,
        tag=tag,
    )


def _read_pdf_metadata(pdf_path: Path) -> dict[str, bytes]:
    """Pull the /Info dictionary's literal strings out of a ReportLab PDF.

    ReportLab serializes Author/Title/Creator/Subject/Keywords as literal
    `(...)` strings inside the `/Info` dictionary. A regex over the raw
    bytes is enough to verify them — we don't need to parse the full
    PDF cross-ref table for a test."""
    data = pdf_path.read_bytes()
    out: dict[str, bytes] = {}
    for tag in (b"Title", b"Author", b"Creator", b"Subject", b"Keywords"):
        m = re.search(b"/" + tag + rb"\s*\(([^)]*)\)", data)
        if m:
            out[tag.decode()] = m.group(1)
    return out


def _pdf_contains(pdf_path: Path, needle: bytes) -> bool:
    """Best-effort substring search over PDF bytes.

    ReportLab serializes short ASCII strings (like our footer text) as
    literal `(mgz-pkmn)` runs inside content streams, so a raw substring
    match is sufficient for the presence checks below — we don't need a
    true text extractor like pypdf."""
    return needle in pdf_path.read_bytes()


def _pdf_has_footer(pdf_path: Path) -> bool:
    """Footer-specific text that is NOT also in PDF metadata.

    `mgz-pkmn` alone appears in `/Author` and `/Creator`, so matching
    it can't distinguish "footer drew correctly" from "metadata wrote
    correctly". The `generated YYYY-MM-DD` prefix and the `p.N` page
    suffix only show up in the footer content stream — with page
    compression disabled at the top of this module they appear as
    literal text in the PDF bytes."""
    data = pdf_path.read_bytes()
    return b"generated " in data and re.search(rb"p\.\d+", data) is not None


class XlsxBrandingTests(unittest.TestCase):
    def test_workbook_properties_name_the_project(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            write_spreadsheet([_row()], out)
            wb = load_workbook(out)
            self.assertEqual(wb.properties.creator, branding.PROJECT_AUTHOR)
            self.assertEqual(wb.properties.title, "cards")
            self.assertIn(branding.PROJECT_NAME, wb.properties.subject or "")
            self.assertIn("mgz-pkmn", wb.properties.keywords or "")

    def test_logo_image_anchored_in_header_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            write_spreadsheet([_row()], out)
            wb = load_workbook(out)
            ws = wb.active
            # _images is internal but openpyxl exposes nothing else; a
            # single embedded image is enough proof the logo was added.
            self.assertGreaterEqual(len(ws._images), 1)

    def test_totals_row_carries_branded_hyperlink(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            write_spreadsheet([_row()], out)
            wb = load_workbook(out)
            ws = wb.active
            found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == branding.PROJECT_NAME and cell.hyperlink:
                        self.assertEqual(cell.hyperlink.target, branding.PROJECT_URL)
                        found = True
            self.assertTrue(found, "expected an mgz-pkmn cell with a project hyperlink")


class BinderBrandingTests(unittest.TestCase):
    def test_standard_layout_metadata_and_footer(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "binder.pdf"
            write_binder_pdf([_row()], out, layout=STANDARD_LAYOUT)
            meta = _read_pdf_metadata(out)
            self.assertEqual(meta.get("Author"), branding.PROJECT_AUTHOR.encode())
            self.assertIn(b"mgz-pkmn", meta.get("Creator", b""))
            self.assertIn(b"mgz-pkmn", meta.get("Subject", b""))
            self.assertIn(b"mgz-pkmn", meta.get("Keywords", b""))
            self.assertTrue(_pdf_has_footer(out))

    def test_condensed_layout_carries_same_branding(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "binder.pdf"
            write_binder_pdf([_row()], out, layout=CONDENSED_LAYOUT)
            meta = _read_pdf_metadata(out)
            self.assertEqual(meta.get("Author"), branding.PROJECT_AUTHOR.encode())
            self.assertTrue(_pdf_has_footer(out))


class ChecklistBrandingTests(unittest.TestCase):
    def test_checklist_metadata_and_footer(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "checklist.pdf"
            write_checklist_pdf([_row()], out)
            meta = _read_pdf_metadata(out)
            self.assertEqual(meta.get("Author"), branding.PROJECT_AUTHOR.encode())
            self.assertIn(b"mgz-pkmn", meta.get("Creator", b""))
            self.assertTrue(_pdf_has_footer(out))


class SetCardsBrandingTests(unittest.TestCase):
    def test_set_cards_metadata_and_footer(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "set-cards.pdf"
            sets = [
                {
                    "id": "base1",
                    "name": "Base",
                    "series": "Base",
                    "total": 102,
                    "printedTotal": 102,
                    "releaseDate": "1999/01/09",
                    "images": {},
                }
            ]
            write_set_cards_pdf(sets, out)
            meta = _read_pdf_metadata(out)
            self.assertEqual(meta.get("Author"), branding.PROJECT_AUTHOR.encode())
            self.assertIn(b"mgz-pkmn", meta.get("Creator", b""))
            self.assertTrue(_pdf_has_footer(out))


class LogoAssetTests(unittest.TestCase):
    def test_bundled_logo_resolves_and_is_a_png(self) -> None:
        data = branding.logo_bytes()
        # PNG magic number — confirms the asset wasn't replaced with the SVG by accident.
        self.assertEqual(data[:8], b"\x89PNG\r\n\x1a\n")


if __name__ == "__main__":
    unittest.main()
