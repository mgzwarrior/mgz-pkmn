"""Tests for the configurable xlsx columns (#262)."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from mgz_pkmn.export_fields import (
    ADJUSTED_COMP_80,
    ADJUSTED_MARKET,
    COMP_80,
    CONDITION,
    MARKET,
    NAME,
    SET,
)
from mgz_pkmn.parser import CardQuery
from mgz_pkmn.pricing import Pricing
from mgz_pkmn.spreadsheet import HEADERS, Row, write_spreadsheet


def _row(market: float | None = 12.5) -> Row:
    return Row(
        query=CardQuery(raw="Charizard", name="Charizard"),
        card={
            "id": "x",
            "name": "Charizard",
            "number": "4",
            "rarity": "Rare Holo",
            "set": {"name": "Base"},
        },
        pricing=Pricing(market=market, variant="holo", source="tcgplayer", url="https://example/x"),
        tag="list-a",
    )


def _write_and_load(fields):
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "cards.xlsx"
        write_spreadsheet([_row()], out, fields=fields)
        return load_workbook(out).active


class DefaultBehaviorTests(unittest.TestCase):
    def test_fields_none_renders_every_header(self) -> None:
        ws = _write_and_load(None)
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header_row, HEADERS)


class FieldFilterTests(unittest.TestCase):
    def test_disabled_column_is_absent_from_header_row(self) -> None:
        ws = _write_and_load(frozenset({NAME}))
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertNotIn("Rarity", header_row)
        self.assertNotIn("Variant", header_row)
        self.assertNotIn("Market", header_row)
        self.assertIn("Name", header_row)
        # Always-on columns stay regardless of the fields filter.
        self.assertIn("Source", header_row)
        self.assertIn("Input", header_row)

    def test_only_enabled_comp_tier_is_written(self) -> None:
        ws = _write_and_load(frozenset({MARKET, COMP_80}))
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("80%", header_row)
        self.assertNotIn("85%", header_row)
        self.assertNotIn("90%", header_row)
        self.assertNotIn("95%", header_row)

    def test_empty_fields_still_renders_always_on_columns(self) -> None:
        ws = _write_and_load(frozenset())
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            header_row,
            ["Source", "Input", "Series", "Database", "eBay Sold (median)", "eBay Active (floor)"],
        )

    def test_set_and_name_values_land_in_the_right_columns(self) -> None:
        ws = _write_and_load(frozenset({NAME, SET}))
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = header_row.index("Name") + 1
        set_col = header_row.index("Set") + 1
        self.assertEqual(ws.cell(row=2, column=name_col).value, "Charizard")
        self.assertEqual(ws.cell(row=2, column=set_col).value, "Base")

    def test_adjusted_condition_columns_render_alongside_raw_market(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            row = _row(market=100.0)
            row.pricing.condition = "LP"
            row.pricing.condition_multiplier = 0.85
            row.pricing.adjusted_market = 85.0
            write_spreadsheet(
                [row],
                out,
                fields=frozenset({CONDITION, MARKET, ADJUSTED_MARKET, COMP_80, ADJUSTED_COMP_80}),
            )
            ws = load_workbook(out).active

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        condition_col = header_row.index("Condition") + 1
        market_col = header_row.index("Market") + 1
        adjusted_market_col = header_row.index("Adjusted Market") + 1
        comp_col = header_row.index("80%") + 1
        adjusted_comp_col = header_row.index("Adjusted 80%") + 1

        self.assertEqual(ws.cell(row=2, column=condition_col).value, "LP")
        self.assertEqual(ws.cell(row=2, column=market_col).value, 100.0)
        self.assertEqual(ws.cell(row=2, column=adjusted_market_col).value, 85.0)
        self.assertEqual(ws.cell(row=2, column=comp_col).value, 80.0)
        self.assertEqual(ws.cell(row=2, column=adjusted_comp_col).value, 68.0)

    def test_pricing_override_replaces_market_and_its_comps(self) -> None:
        """A manual override (#266) renders in place of market — the raw
        Market column and its 80/85/90/95% comps all use the override
        instead of the source's live price."""
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            row = _row(market=100.0)
            row.pricing.pricing_override = 12.0
            write_spreadsheet([row], out, fields=frozenset({MARKET, COMP_80}))
            ws = load_workbook(out).active

        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        market_col = header_row.index("Market") + 1
        comp_col = header_row.index("80%") + 1

        self.assertEqual(ws.cell(row=2, column=market_col).value, 12.0)
        self.assertEqual(ws.cell(row=2, column=comp_col).value, 9.6)


if __name__ == "__main__":
    unittest.main()
