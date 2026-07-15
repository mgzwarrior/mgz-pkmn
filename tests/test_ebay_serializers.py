"""Tests for the eBay comp signals on the output serializers (#423).

Covers the `summarize_ebay_comps` reducer and the three serializer surfaces
the issue names — xlsx columns, the JSON report rows payload, and (via the
shared `_pricing_to_dict`) the wire shape. The acceptance contract is:
existing fields unchanged when no eBay data is present, the new fields /
columns populated when it is.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from mgz_pkmn.parser import CardQuery
from mgz_pkmn.pricing import Pricing, summarize_ebay_comps
from mgz_pkmn.report import build_json_report
from mgz_pkmn.spreadsheet import HEADERS, Row, write_spreadsheet


def _row(pricing: Pricing, tag: str = "list-a") -> Row:
    return Row(
        query=CardQuery(raw="Charizard", name="Charizard"),
        card={"id": "x", "name": "Charizard", "number": "4", "set": {"name": "Base"}},
        pricing=pricing,
        tag=tag,
    )


class SummarizeEbayCompsTests(unittest.TestCase):
    def test_median_sold_and_active_floor(self) -> None:
        comps = [
            Pricing(market=230.0, source="ebay_sold"),
            Pricing(market=250.0, source="ebay_sold"),
            Pricing(market=210.0, source="ebay_sold"),
            Pricing(market=199.99, source="ebay_active"),
            Pricing(market=240.0, source="ebay_active"),
        ]
        median_sold, active_floor = summarize_ebay_comps(comps)
        self.assertEqual(median_sold, 230.0)  # median of 210/230/250
        self.assertEqual(active_floor, 199.99)  # cheapest active

    def test_each_tier_independent(self) -> None:
        # Only active comps present → sold stays None, and vice versa.
        self.assertEqual(
            summarize_ebay_comps([Pricing(market=10.0, source="ebay_active")]),
            (None, 10.0),
        )
        self.assertEqual(
            summarize_ebay_comps([Pricing(market=10.0, source="ebay_sold")]),
            (10.0, None),
        )

    def test_empty_and_unpriced_comps_yield_none(self) -> None:
        self.assertEqual(summarize_ebay_comps([]), (None, None))
        self.assertEqual(
            summarize_ebay_comps([Pricing(market=None, source="ebay_sold")]),
            (None, None),
        )

    def test_ignores_non_ebay_sources(self) -> None:
        self.assertEqual(
            summarize_ebay_comps([Pricing(market=5.0, source="tcgplayer")]),
            (None, None),
        )


class PricingDefaultsTests(unittest.TestCase):
    def test_ebay_fields_default_to_none(self) -> None:
        p = Pricing(market=12.5)
        self.assertIsNone(p.ebay_sold_median)
        self.assertIsNone(p.ebay_active_floor)


class XlsxColumnTests(unittest.TestCase):
    def test_headers_carry_ebay_columns_before_price_source(self) -> None:
        self.assertEqual(
            HEADERS[21:25],
            ["eBay Sold (median)", "eBay Active (floor)", "Price Source", "Listing URL"],
        )

    def _write_and_load(self, pricing: Pricing):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "cards.xlsx"
            write_spreadsheet([_row(pricing)], out)
            return load_workbook(out).active

    def test_ebay_cells_populated_when_present(self) -> None:
        ws = self._write_and_load(
            Pricing(
                market=250.0, source="ebay_active", ebay_sold_median=230.0, ebay_active_floor=199.99
            )
        )
        # Row 2, columns V (22) / W (23) carry the aggregates; X (24) the source.
        self.assertEqual(ws.cell(row=2, column=22).value, 230.0)
        self.assertEqual(ws.cell(row=2, column=23).value, 199.99)
        self.assertEqual(ws.cell(row=2, column=24).value, "ebay_active")

    def test_ebay_cells_dash_when_absent(self) -> None:
        ws = self._write_and_load(Pricing(market=12.5, source="tcgplayer"))
        self.assertEqual(ws.cell(row=2, column=22).value, "—")
        self.assertEqual(ws.cell(row=2, column=23).value, "—")
        self.assertEqual(ws.cell(row=2, column=24).value, "tcgplayer")


class JsonReportTests(unittest.TestCase):
    def _rows_payload(self, pricing: Pricing) -> dict:
        payload = build_json_report(rows=[_row(pricing)], counters={}, input_lines=1, elapsed=1.0)
        return payload["rows"][0]

    def test_rows_carry_ebay_signals_when_present(self) -> None:
        row = self._rows_payload(
            Pricing(market=250.0, ebay_sold_median=230.0, ebay_active_floor=199.99)
        )
        self.assertEqual(row["ebay_sold_median"], 230.0)
        self.assertEqual(row["ebay_active_floor"], 199.99)

    def test_rows_carry_null_ebay_signals_when_absent(self) -> None:
        row = self._rows_payload(Pricing(market=12.5))
        self.assertIsNone(row["ebay_sold_median"])
        self.assertIsNone(row["ebay_active_floor"])


if __name__ == "__main__":
    unittest.main()
